
# core.py
# -*- coding: utf-8 -*-
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
import calendar
from typing import List, Optional, Iterable, Tuple, Dict, Any, List

# -------------------- CONSTANTES CENTRALIZADAS --------------------
# SMI vigente (RD 87/2025), prórroga transitoria a 2026 hasta nuevo RD
SMI_VIGENTE_ANUAL_14P   = 16_576.00   # €  (1.184 €/mes * 14)
SMI_VIGENTE_MENSUAL_14P = 1_184.00    # €
# Revalorización de pensión (objetivo IPC BdE) aplicada 1 de enero
PENSION_REVAL_IPC_OBJ   = 0.02        # 2%
# Compensación de empresa en SEPE (aportación trabajador a SS)
PCT_COMP_SEPE_SS_EMPRESA = 0.047      # 4,70%
# Rutas ficheros de datos (bases, coeficientes, evolución pensión máxima, etc.)
RUTA_BASES_COTIZACION = "data/inputs/bases_cotizacion.txt"  # CSV con columnas 'fecha' (YYYY-MM-01) y 'base' (float)
RUTA_BASES_OK = "data/inputs/bases_cotizacion_ok.txt"  # CSV con columnas 'Año', 'Empresa', 'Enero', ..., 'Diciembre' (salida de txt2bases_csv.py)
RUTA_TXT_BASES="data/inputs/Informe Bases Cotización Online.txt"  # Ruta del txt original descargado de la Seguridad Social (sin formato tabular)
RUTA_INCREMENTO_BASES_REGULADORAS = "app/static/incremento_bases_reguladoras.txt"  # CSV con columnas 'fecha' (YYYY-MM-01) y 'pct_incremento' (float)
RUTA_COEFICIENTES_JAI = "app/static/coeficientes_jubilacion_anticipada_involuntaria.txt"  # CSV con columnas 'fecha' (YYYY-MM-01) y 'coeficiente' (float)
RUTA_COEFICIENTES_JAV = "app/static/coeficientes_jubilacion_anticipada_voluntaria.txt"  # CSV con columnas 'fecha' (YYYY-MM-01) y 'coeficiente' (float)
RUTA_EVOLUCION_PENSION_MAXIMA = "app/static/evolucion_pension_maxima.txt"  # CSV con columnas 'fecha' (YYYY-MM-01) y 'pension_max_mensual' (float)
RUTA_BASES_MAXIMAS = "app/static/bases_min_max.txt"  # CSV con columnas 'fecha' (YYYY-MM-01) y 'base_max' (float)
RUTA_BRECHA_GENERO = "app/static/Brecha_Genero.txt"  # CSV con columnas Año;Importe Mensual (1 hijo);% Subida (Est.)
RUTA_EXCEL_RESUMEN_JUBILACION = "data/outputs/Resumen_Calculo_Jubilacion.xlsx"  # Excel de salida con resumen y tablas
RUTA_TABLA_IPC = "app/static/Tabla_IPC.txt"  # CSV con columnas 'Año', 'IPC' (índice anual), 'Estado'
RUTA_PRESTACION_MINMAX: str = "app/static/prestacion_contributiva_espana.txt"
RUTA_DURACION_PRESTACION: str = "app/static/duracion_prestacion_contributiva.txt"
RUTA_EV = "app/static/Esperanza_Vida.txt"  # Formato: Año;Hombres;Mujeres
EXCEL_SALIDA_PATH = "data/outputs/Resumen_Rentas.xlsx"  # Excel de salida con resultados detallados iterativos
UMBRAL_ULTIMOS_MESES: int = 12
UMBRAL_MIN_MATCH: int = 12
EPS_TOLERANCIA_MAXIMA: float = 0.5
# --- Parámetros CESS (empresa) ---
PCT_CESS_EMPRESA = 0.2830          # 28,30%
COEF_REDUCTOR_CESS = 0.94          # 0,94
PCT_MEI_2023 = 0.0090              # 0,90% (desde 2023)
MEI_START = datetime(2023, 1, 1)   # umbral MEI

# -------------------- HELPERS DE FECHA --------------------
def _month_start(d: datetime) -> datetime:
    return datetime(d.year, d.month, 1)

def _add_months(d: datetime, m: int) -> datetime:
    return d + relativedelta(months=m)

def _month_end(d: datetime) -> datetime:
    ms = _month_start(d)
    last_day = calendar.monthrange(ms.year, ms.month)[1]
    return datetime(ms.year, ms.month, last_day)

def _nearest_month_start(d: datetime) -> datetime:
    """Devuelve el día 1 más cercano a la fecha d (mismo mes o siguiente)."""
    prev_start = _month_start(d)
    next_start = _add_months(prev_start, 1)
    diff_prev = abs((d - prev_start).days)
    diff_next = abs((next_start - d).days)
    return prev_start if diff_prev <= diff_next else next_start

def _days_between(d1: datetime, d2: datetime) -> int:
    """
    Días naturales entre d1 y d2 (sin +1 para no sobrecontar). Si d2 < d1, 0.
    """
    return max(0, (d2 - d1).days)

# -------------------- LECTURA DE TABLAS --------------------
def _read_table(path, sep=';', decimal=','):
    try:
        return pd.read_csv(path, sep=sep, decimal=decimal)
    except Exception:
        return pd.read_csv(path, sep=sep, decimal='.')

# -------------------- HELPERS DE BASES --------------------
def _get_base_mensual_para_mes(df_mensual: pd.DataFrame, mes_fecha: datetime) -> float:
    """Devuelve la base para el mes; si falta, usa la última conocida hacia atrás."""
    mes = _month_start(pd.to_datetime(mes_fecha))
    s = df_mensual.copy()
    s['fecha'] = pd.to_datetime(s['fecha']).dt.to_period('M').dt.to_timestamp()
    s = s.sort_values('fecha').set_index('fecha')['base']
    if mes in s.index:
        return float(s.loc[mes])
    prev = s[s.index < mes]
    return float(prev.iloc[-1]) if not prev.empty else 0.0

'''
def _media_6_meses_previos_para_paro(inicio_prest: datetime, df_mensual: pd.DataFrame) -> float:
    """Media de las bases de los 6 meses inmediatamente anteriores a 'inicio_prest'."""
    vals = []
    for i in range(1, 7):
        mes_ref = _add_months(_month_start(inicio_prest), -i)
        vals.append(_get_base_mensual_para_mes(df_mensual, mes_ref))
    return float(np.mean(vals)) if vals else 0.0
'''

def _media_180_dias_previos_para_paro(fecha_inicio: datetime, df_mensual: pd.DataFrame) -> float:
    """
    Media diaria de los últimos 180 días previos a 'fecha_inicio' (baja);
    base diaria = base mensual / 30. Devuelve base mensual equivalente.
    """
    ventana_fin = fecha_inicio - timedelta(days=1)
    ventana_ini = fecha_inicio - timedelta(days=180)
    s = df_mensual.set_index('fecha')['base'].copy()
    s.index = s.index.map(_month_start)

    cursor = _month_start(ventana_ini)
    suma_diaria = 0.0
    dias_total = 0
    while cursor <= ventana_fin:
        inicio_mes = max(cursor, ventana_ini)
        fin_mes = min(_month_end(cursor), ventana_fin)
        dias_solape = (fin_mes - inicio_mes).days + 1
        # base del mes (fallback a última hacia atrás si faltara)
        if cursor in s.index:
            base_mes = float(s.loc[cursor])
        else:
            prev = s[s.index < cursor]
            base_mes = float(prev.iloc[-1]) if not prev.empty else 0.0
        suma_diaria += (base_mes / 30.0) * dias_solape
        dias_total += dias_solape
        cursor = _add_months(cursor, 1)

    if dias_total == 0:
        raise ValueError("No hay días válidos en los 180 previos a la prestación.")
    return (suma_diaria / dias_total) * 30.0

# -------------------- VALIDACIONES --------------------
def _validar_entradas_jubilacion(
    *,
    fecha_nacimiento: datetime,
    fecha_baja_ere_despido: datetime,
    fecha_jubilacion_anticipada: datetime,
    causa_involuntaria: bool,
    df_bases_in: pd.DataFrame = None
) -> None:
    """
    1) fecha_nacimiento <= 31-12-1971
    2) fecha_baja_ere_despido >= 01-03-2026
    3) fecha_baja_ere_despido día 1 de mes
    4) fecha_jubilacion_anticipada >= (edad_65 - (4 si involuntaria | 2 si voluntaria))
    """
    errores: List[str] = []

    if df_bases_in is not None:
        if not isinstance(df_bases_in, pd.DataFrame):
            errores.append("- 'df_bases_in' debe ser un DataFrame de pandas.")
        elif not {'Año', 'Empresa'}.issubset(df_bases_in.columns):
            errores.append("- 'df_bases_in' debe contener las columnas 'Año' y 'Empresa'.")
    else:
        errores.append("- 'df_bases_in' es requerido para el cálculo de jubilación anticipada. No puede ser None.")
    
    tope_nacimiento = datetime(1971, 12, 31)
    if fecha_nacimiento > tope_nacimiento:
        errores.append(f"- 'fecha_nacimiento' debe ser <= {tope_nacimiento.date()}.")

    minimo_baja = datetime(2026, 3, 1)
    if fecha_baja_ere_despido < minimo_baja:
        errores.append(f"- 'fecha_baja_ere_despido' debe ser >= {minimo_baja.date()}.")

    if fecha_baja_ere_despido.day != 1:
        errores.append("- 'fecha_baja_ere_despido' debe ser el día 1 de algún mes.")

    edad_65 = fecha_nacimiento + relativedelta(years=65)
    anticipo_max_anios = 4 if causa_involuntaria else 2
    fecha_min_anticipada = edad_65 - relativedelta(years=anticipo_max_anios)
    if fecha_jubilacion_anticipada < fecha_min_anticipada:
        etiqueta = "involuntaria" if causa_involuntaria else "voluntaria"
        errores.append(
            f"- 'fecha_jubilacion_anticipada' ({fecha_jubilacion_anticipada.date()}) "
            f"debe ser >= {fecha_min_anticipada.date()} (anticipación máxima {anticipo_max_anios} años, causa {etiqueta})."
        )

    if errores:
        raise ValueError("VALIDACIÓN DE ENTRADA (Jubilación anticipada) FALLIDA:\n" + "\n".join(errores))

def _validate_inputs_rentas(
    *,
    fecha_nacimiento: datetime,
    fecha_baja: datetime,
    salario_fijo_anual: float,
    bonus_target_anual: float,
    complementos: float,
    pct_renta_hasta_63: float,
    pct_renta_hasta_65: float,
    pct_reval_desde_63: float,
    num_hijos: int
) -> None:
    """
    1) fecha_nacimiento <= 31-12-1971
    2) fecha_baja >= 01-03-2026
    3) fecha_baja día 1
    4) salario_fijo_anual > SMI vigente (hardcodeado)
    5) bonus_target_anual <= salario_fijo_anual + complementos
    6) porcentajes en [0,1]
    7) num_hijos >= 0
    """
    errores: List[str] = []

    min_nac = datetime(1971, 12, 31)
    if fecha_nacimiento > min_nac:
        errores.append(f"- 'fecha_nacimiento' debe ser <= {min_nac.date()}.")

    min_baja = datetime(2026, 3, 1)
    if fecha_baja < min_baja:
        errores.append(f"- 'fecha_baja' debe ser >= {min_baja.date()}.")

    if fecha_baja.day != 1:
        errores.append("- 'fecha_baja' debe ser el día 1 de algún mes.")

    if salario_fijo_anual <= SMI_VIGENTE_ANUAL_14P:
        errores.append(f"- 'salario_fijo_anual' ({salario_fijo_anual:,.2f}) debe ser > SMI anual vigente ({SMI_VIGENTE_ANUAL_14P:,.2f}).")

    if bonus_target_anual > (salario_fijo_anual + complementos):
        errores.append(f"- 'bonus_target_anual' ({bonus_target_anual:,.2f}) no puede exceder 'salario_fijo_anual + complementos' ({(salario_fijo_anual + complementos):,.2f}).")

    for nombre, val in [
        ("pct_renta_hasta_63", pct_renta_hasta_63),
        ("pct_renta_hasta_65", pct_renta_hasta_65),
        ("pct_reval_desde_63", pct_reval_desde_63),
    ]:
        if not (0.0 <= val <= 1.0):
            errores.append(f"- '{nombre}' debe estar entre 0 y 1 (incluidos). Valor actual: {val}")

    if num_hijos < 0:
        errores.append("- 'num_hijos' debe ser >= 0.")

    if errores:
        raise ValueError("VALIDACIÓN DE ENTRADA FALLIDA:\n" + "\n".join(errores))

# Helpers de escritura en Excel (2 y 3 columnas)

def _append_key_values_to_sheet(
    xlsx_path: str,
    sheet_name: str,
    block_title: str,
    kv_pairs: Iterable[Tuple[str, object]],
) -> None:
    """
    Apende un bloque de 2 columnas (Parametro, Valor) al final de una hoja Excel,
    sin borrar contenido previo. Crea la hoja si no existe.
    """
    from openpyxl import load_workbook
    import pandas as pd

    try:
        wb = load_workbook(xlsx_path)
    except FileNotFoundError:
        with pd.ExcelWriter(xlsx_path, engine="openpyxl") as w:
            pd.DataFrame().to_excel(w, sheet_name=sheet_name, index=False)
        wb = load_workbook(xlsx_path)

    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(title=sheet_name)

    row = ws.max_row or 0
    if row == 1 and all(c.value is None for c in ws[1]):
        row = 0
    start_row = row + 2 if row > 0 else 1

    ws.cell(row=start_row,   column=1, value=block_title)
    ws.cell(row=start_row+1, column=1, value="Parametro")
    ws.cell(row=start_row+1, column=2, value="Valor")

    r = start_row + 2
    for k, v in kv_pairs:
        ws.cell(row=r, column=1, value=k)
        ws.cell(row=r, column=2, value=v)
        r += 1

    wb.save(xlsx_path)


def _append_triplets_to_sheet(
    xlsx_path: str,
    sheet_name: str,
    block_title: str,
    triplets: Iterable[Tuple[str, str, object]],
) -> None:
    """
    Apende un bloque de 3 columnas (Grupo, Campo, Valor) al final de una hoja Excel,
    sin borrar contenido previo. Crea la hoja si no existe.
    """
    from openpyxl import load_workbook
    import pandas as pd

    rows = []
    for i, t in enumerate(triplets, start=1):
        if not isinstance(t, (list, tuple)) or len(t) != 3:
            raise ValueError(f"Entrada #{i}: se esperaba tripleta (grupo, campo, valor). Recibido: {repr(t)}")
        g, c, v = t
        rows.append((str(g), str(c), v))
    if not rows:
        return

    try:
        wb = load_workbook(xlsx_path)
    except FileNotFoundError:
        with pd.ExcelWriter(xlsx_path, engine="openpyxl") as w:
            pd.DataFrame().to_excel(w, sheet_name=sheet_name, index=False)
        wb = load_workbook(xlsx_path)

    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(title=sheet_name)

    row = ws.max_row or 0
    if row == 1 and all(c.value is None for c in ws[1]):
        row = 0
    start_row = row + 2 if row > 0 else 1

    ws.cell(row=start_row,   column=1, value=block_title)
    ws.cell(row=start_row+1, column=1, value="Grupo")
    ws.cell(row=start_row+1, column=2, value="Campo")
    ws.cell(row=start_row+1, column=3, value="Valor")

    r = start_row + 2
    for g, c, v in rows:
        ws.cell(row=r, column=1, value=g)
        ws.cell(row=r, column=2, value=c)
        ws.cell(row=r, column=3, value=v)
        r += 1

    wb.save(xlsx_path)

# ---------------------- Utilidades de Excel (append) ---------------------- #
def _append_rows_to_excel(xlsx_path: str, sheet_name: str, rows: List[Dict[str, Any]]) -> None:
    """Apende filas (lista de diccionarios homogéneos) a un Excel/hoja.
    Crea el fichero/hoja si no existen. Mantiene encabezados de la primera escritura.
    """
    df_new = pd.DataFrame(rows)
    try:
        with pd.ExcelWriter(xlsx_path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
            # Si la hoja existe: escribimos sin cabecera a partir de la última fila
            startrow = writer.sheets[sheet_name].max_row if sheet_name in writer.sheets else 0
            df_new.to_excel(writer, sheet_name=sheet_name, index=False, header=(startrow == 0), startrow=startrow)
    except FileNotFoundError:
        # Crear desde cero
        with pd.ExcelWriter(xlsx_path, engine="openpyxl", mode="w") as writer:
            df_new.to_excel(writer, sheet_name=sheet_name, index=False)
    except ValueError:
        # Caso típico: la hoja no existe aún en modo append -> creamos la hoja
        with pd.ExcelWriter(xlsx_path, engine="openpyxl", mode="a") as writer:
            df_new.to_excel(writer, sheet_name=sheet_name, index=False)


def _log_kv_rows(iter_idx: int, fecha_jub_ant: datetime, paso: str, kv: Dict[str, Any]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for k, v in kv.items():
        # Evitar DataFrames y objetos complejos; sólo escalares/strings
        if isinstance(v, (pd.DataFrame, pd.Series)):
            continue
        rows.append(
            {
                "iter": iter_idx,
                "fecha_jubilacion_anticipada": fecha_jub_ant.strftime("%Y-%m-%d"),
                "paso": paso,
                "clave": str(k),
                "valor": v,
            }
        )
    return rows

# NUEVO: configuración de escenario
from dataclasses import dataclass
from typing import Literal, Optional

# -----------------------------Definición de Escenario---------------------------------- #
@dataclass(frozen=True)
class Escenario:
    modalidad: Literal["ERE", "PSI", "OTRO"] = "ERE"
    tiene_prestacion_desempleo: bool = True
    meses_prestacion: int = 24
    causa_involuntaria: bool = True  # PSI fuerza False

    @staticmethod
    def from_modalidad(modalidad: str) -> "Escenario":
        m = (modalidad or "ERE").upper()
        if m == "PSI":
            return Escenario(modalidad="PSI",
                             tiene_prestacion_desempleo=False,
                             meses_prestacion=0,
                             causa_involuntaria=False)
        elif m == "ERE":
            return Escenario()
        else:
            # 'OTRO' deja valores neutros; el llamador podrá ajustarlos
            return Escenario(modalidad="OTRO", tiene_prestacion_desempleo=False,
                             meses_prestacion=0, causa_involuntaria=False)